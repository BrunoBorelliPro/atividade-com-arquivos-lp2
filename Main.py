from openpyxl import Workbook, load_workbook
from Funcionario import Funcionario #Importa a classe "Funcionario"
#Essa parte do código tenta ler a planilha .xlsx e caso ocorra um erro ele cria uma planilha nova com os sheets "Funcionarios" e "Despesas"
try:
    wb = load_workbook("AtividadeComArquivos.xlsx")
except:
    wb = Workbook()
    wb.create_sheet(title="Funcionarios",index=0)
    wb.create_sheet(title="Despesas",index=1)

# Menu
while True:
    print("1 - Adicionar funcionário")
    print("2 - Listar Funcionários")
    print("0 - Sair")
    opt = input("Selecione a opção: ")
    if opt == "1":
        id = input("Insira a Identificação do funcionário: ")
        nome = input("Insira o nome do funcionário: ")
        try:
            funcionario = Funcionario(id,nome)
            funcionario.Adiciona_funcionario(wb)
        except:
            print("Erro ao adicionar um funcionario")
    elif opt == "0":
        print("Encerrado!")
        break
    elif opt == "2":
        Funcionario.Listar_Funcionarios(wb)
    else:
        print("Opção inválida!")
        print("")

